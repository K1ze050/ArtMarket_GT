import tkinter as tk
from tkinter import messagebox, ttk
from datetime import datetime
from openpyxl import load_workbook
import os

# Importar las clases del programa original
# IMPORTANTE: Asegúrate de que este archivo esté en la misma carpeta que tu programa original
# y que tu programa original se llame exactamente como lo importas aquí

# Configuración de archivos Excel (debe coincidir con tu programa original)
ESCRITORIO = os.path.join(os.path.expanduser("~"), "Desktop")
ARCHIVO_EXCEL = os.path.join(ESCRITORIO, "base_datos.xlsx")
HOJA_INVENTARIO = "Inventario"
HOJA_TRABAJOS = "Trabajos"

# Lista de productos y trabajos válidos
PRODUCTOS_VALIDOS = ["playera", "taza", "papel sublimado", "vidrio", "papel impresión", "vinil"]
TRABAJOS_VALIDOS = ["corte eléctrico en vinil adhesivo", "sublimado"]


class InterfazGrafica:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Gestión - Inventario y Trabajos")
        self.root.geometry("600x500")
        self.root.resizable(False, False)
        
        # Configurar estilo
        self.configurar_estilo()
        
        # Crear la interfaz principal
        self.crear_ventana_principal()
    
    def configurar_estilo(self):
        """Configura los estilos de la interfaz"""
        self.root.configure(bg="#f0f0f0")
    
    def crear_ventana_principal(self):
        """Crea la ventana principal con los botones del menú"""
        # Título
        titulo = tk.Label(
            self.root,
            text="SISTEMA DE GESTIÓN",
            font=("Arial", 24, "bold"),
            bg="#f0f0f0",
            fg="#333333"
        )
        titulo.pack(pady=30)
        
        subtitulo = tk.Label(
            self.root,
            text="Inventario y Trabajos",
            font=("Arial", 12),
            bg="#f0f0f0",
            fg="#666666"
        )
        subtitulo.pack(pady=5)
        
        # Frame para los botones
        frame_botones = tk.Frame(self.root, bg="#f0f0f0")
        frame_botones.pack(pady=30)
        
        # Botones del menú
        botones = [
            ("Registrar Trabajo", self.ventana_registrar_trabajo),
            ("Agregar Inventario", self.ventana_agregar_inventario),
            ("Mostrar Almacén", self.ventana_mostrar_almacen),
            ("Buscar Producto", self.ventana_buscar_producto),
            ("Ver Trabajos Registrados", self.ventana_ver_trabajos),
            ("Salir", self.salir)
        ]
        
        for texto, comando in botones:
            btn = tk.Button(
                frame_botones,
                text=texto,
                command=comando,
                width=30,
                height=2,
                font=("Arial", 11),
                bg="#4CAF50" if texto != "Salir" else "#f44336",
                fg="white",
                cursor="hand2",
                relief=tk.RAISED,
                bd=3
            )
            btn.pack(pady=8)
    
    def ventana_registrar_trabajo(self):
        """Ventana para registrar un nuevo trabajo"""
        ventana = tk.Toplevel(self.root)
        ventana.title("Registrar Trabajo")
        ventana.geometry("500x450")
        ventana.resizable(False, False)
        
        # Título
        tk.Label(
            ventana,
            text="Registrar Nuevo Trabajo",
            font=("Arial", 16, "bold")
        ).pack(pady=20)
        
        # Frame para el formulario
        frame = tk.Frame(ventana)
        frame.pack(pady=10, padx=20)
        
        # Campo Cliente
        tk.Label(frame, text="Cliente:", font=("Arial", 10)).grid(row=0, column=0, sticky="w", pady=10)
        entry_cliente = tk.Entry(frame, width=40, font=("Arial", 10))
        entry_cliente.grid(row=0, column=1, pady=10)
        
        # Campo Tipo de Trabajo
        tk.Label(frame, text="Tipo de Trabajo:", font=("Arial", 10)).grid(row=1, column=0, sticky="w", pady=10)
        combo_trabajo = ttk.Combobox(
            frame,
            values=[t.capitalize() for t in TRABAJOS_VALIDOS],
            width=38,
            font=("Arial", 10),
            state="readonly"
        )
        combo_trabajo.grid(row=1, column=1, pady=10)
        
        # Campo Fecha de Entrega
        tk.Label(frame, text="Fecha (dd-mm-yyyy):", font=("Arial", 10)).grid(row=2, column=0, sticky="w", pady=10)
        entry_fecha = tk.Entry(frame, width=40, font=("Arial", 10))
        entry_fecha.grid(row=2, column=1, pady=10)
        entry_fecha.insert(0, datetime.now().strftime("%d-%m-%Y"))
        
        # Botón Registrar
        def registrar():
            cliente = entry_cliente.get().strip()
            trabajo = combo_trabajo.get().lower()
            fecha = entry_fecha.get().strip()
            
            if not cliente or not trabajo or not fecha:
                messagebox.showerror("Error", "Todos los campos son obligatorios")
                return
            
            # Validar fecha
            try:
                datetime.strptime(fecha, "%d-%m-%Y")
            except:
                messagebox.showerror("Error", "Formato de fecha inválido. Use dd-mm-yyyy")
                return
            
            # Procesar según el tipo de trabajo
            if trabajo == "corte eléctrico en vinil adhesivo":
                self.procesar_corte_vinil(cliente, trabajo, fecha, ventana)
            elif trabajo == "sublimado":
                self.procesar_sublimado(cliente, trabajo, fecha, ventana)
        
        tk.Button(
            ventana,
            text="Registrar y Descontar Materiales",
            command=registrar,
            width=30,
            height=2,
            font=("Arial", 11, "bold"),
            bg="#4CAF50",
            fg="white",
            cursor="hand2"
        ).pack(pady=20)
    
    def procesar_corte_vinil(self, cliente, trabajo, fecha, ventana_padre):
        """Procesa el descuento de vinil"""
        ventana = tk.Toplevel(ventana_padre)
        ventana.title("Cantidad de Vinil")
        ventana.geometry("400x200")
        
        tk.Label(
            ventana,
            text="Ingrese la cantidad de hojas de vinil:",
            font=("Arial", 12)
        ).pack(pady=20)
        
        entry_cantidad = tk.Entry(ventana, width=20, font=("Arial", 12))
        entry_cantidad.pack(pady=10)
        entry_cantidad.focus()
        
        def confirmar():
            try:
                cantidad = float(entry_cantidad.get())
                
                if cantidad <= 0:
                    messagebox.showerror("Error", "La cantidad debe ser mayor a cero")
                    return
                
                # Verificar y descontar
                wb = load_workbook(ARCHIVO_EXCEL)
                ws = wb[HOJA_INVENTARIO]
                
                vinil_disponible = 0
                fila_vinil = None
                
                for row in range(2, ws.max_row + 1):
                    producto = ws.cell(row, 1).value
                    if producto and producto.lower() == "vinil":
                        vinil_disponible = ws.cell(row, 2).value or 0
                        fila_vinil = row
                        break
                
                if fila_vinil is None:
                    messagebox.showerror("Error", "El producto vinil no está en el inventario")
                    return
                
                if vinil_disponible < cantidad:
                    messagebox.showerror(
                        "Error",
                        f"Materiales insuficientes\nDisponible: {vinil_disponible}\nSolicitado: {cantidad}"
                    )
                    return
                
                # Descontar
                nueva_cantidad = vinil_disponible - cantidad
                ws.cell(fila_vinil, 2, nueva_cantidad)
                
                # Guardar trabajo
                ws_trabajos = wb[HOJA_TRABAJOS]
                ws_trabajos.append([cliente, trabajo, fecha])
                
                wb.save(ARCHIVO_EXCEL)
                
                messagebox.showinfo(
                    "Éxito",
                    f"Trabajo registrado exitosamente\n\nMateriales descontados:\nVinil: {cantidad} hojas\nVinil restante: {nueva_cantidad}"
                )
                
                ventana.destroy()
                ventana_padre.destroy()
                
            except ValueError:
                messagebox.showerror("Error", "Ingrese un número válido")
            except Exception as e:
                messagebox.showerror("Error", f"Error al procesar: {str(e)}")
        
        tk.Button(
            ventana,
            text="Confirmar",
            command=confirmar,
            width=15,
            font=("Arial", 11),
            bg="#4CAF50",
            fg="white",
            cursor="hand2"
        ).pack(pady=10)
    
    def procesar_sublimado(self, cliente, trabajo, fecha, ventana_padre):
        """Procesa el descuento de materiales para sublimado"""
        ventana = tk.Toplevel(ventana_padre)
        ventana.title("Sublimado - Seleccionar Producto")
        ventana.geometry("400x300")
        
        tk.Label(
            ventana,
            text="Seleccione el producto para sublimado:",
            font=("Arial", 12, "bold")
        ).pack(pady=20)
        
        productos_sublimado = ["playera", "taza", "vidrio"]
        
        combo_producto = ttk.Combobox(
            ventana,
            values=[p.capitalize() for p in productos_sublimado],
            width=20,
            font=("Arial", 11),
            state="readonly"
        )
        combo_producto.pack(pady=10)
        combo_producto.current(0)
        
        tk.Label(ventana, text="Cantidad:", font=("Arial", 12)).pack(pady=10)
        
        entry_cantidad = tk.Entry(ventana, width=20, font=("Arial", 12))
        entry_cantidad.pack(pady=10)
        
        def confirmar():
            try:
                producto_elegido = combo_producto.get().lower()
                cantidad = float(entry_cantidad.get())
                
                if cantidad <= 0:
                    messagebox.showerror("Error", "La cantidad debe ser mayor a cero")
                    return
                
                # Verificar y descontar
                wb = load_workbook(ARCHIVO_EXCEL)
                ws = wb[HOJA_INVENTARIO]
                
                producto_disponible = 0
                papel_disponible = 0
                fila_producto = None
                fila_papel = None
                
                for row in range(2, ws.max_row + 1):
                    producto = ws.cell(row, 1).value
                    if producto:
                        if producto.lower() == producto_elegido:
                            producto_disponible = ws.cell(row, 2).value or 0
                            fila_producto = row
                        elif producto.lower() == "papel impresión":
                            papel_disponible = ws.cell(row, 2).value or 0
                            fila_papel = row
                
                if fila_producto is None:
                    messagebox.showerror("Error", f"El producto {producto_elegido} no está en el inventario")
                    return
                
                if fila_papel is None:
                    messagebox.showerror("Error", "El papel impresión no está en el inventario")
                    return
                
                if producto_disponible < cantidad:
                    messagebox.showerror(
                        "Error",
                        f"Materiales insuficientes\n{producto_elegido.capitalize()} disponible: {producto_disponible}\nSolicitado: {cantidad}"
                    )
                    return
                
                if papel_disponible < cantidad:
                    messagebox.showerror(
                        "Error",
                        f"Materiales insuficientes\nPapel impresión disponible: {papel_disponible}\nSolicitado: {cantidad}"
                    )
                    return
                
                # Descontar
                nueva_cantidad_producto = producto_disponible - cantidad
                nueva_cantidad_papel = papel_disponible - cantidad
                
                ws.cell(fila_producto, 2, nueva_cantidad_producto)
                ws.cell(fila_papel, 2, nueva_cantidad_papel)
                
                # Guardar trabajo
                ws_trabajos = wb[HOJA_TRABAJOS]
                ws_trabajos.append([cliente, trabajo, fecha])
                
                wb.save(ARCHIVO_EXCEL)
                
                messagebox.showinfo(
                    "Éxito",
                    f"Trabajo registrado exitosamente\n\nMateriales descontados:\n{producto_elegido.capitalize()}: {cantidad} unidades (restante: {nueva_cantidad_producto})\nPapel impresión: {cantidad} hojas (restante: {nueva_cantidad_papel})"
                )
                
                ventana.destroy()
                ventana_padre.destroy()
                
            except ValueError:
                messagebox.showerror("Error", "Ingrese un número válido")
            except Exception as e:
                messagebox.showerror("Error", f"Error al procesar: {str(e)}")
        
        tk.Button(
            ventana,
            text="Confirmar",
            command=confirmar,
            width=15,
            font=("Arial", 11),
            bg="#4CAF50",
            fg="white",
            cursor="hand2"
        ).pack(pady=20)
    
    def ventana_agregar_inventario(self):
        """Ventana para agregar productos al inventario"""
        ventana = tk.Toplevel(self.root)
        ventana.title("Agregar Inventario")
        ventana.geometry("450x300")
        ventana.resizable(False, False)
        
        tk.Label(
            ventana,
            text="Agregar Producto al Inventario",
            font=("Arial", 16, "bold")
        ).pack(pady=20)
        
        frame = tk.Frame(ventana)
        frame.pack(pady=10, padx=20)
        
        tk.Label(frame, text="Producto:", font=("Arial", 10)).grid(row=0, column=0, sticky="w", pady=10)
        combo_producto = ttk.Combobox(
            frame,
            values=[p.capitalize() for p in PRODUCTOS_VALIDOS],
            width=30,
            font=("Arial", 10),
            state="readonly"
        )
        combo_producto.grid(row=0, column=1, pady=10)
        
        tk.Label(frame, text="Cantidad:", font=("Arial", 10)).grid(row=1, column=0, sticky="w", pady=10)
        entry_cantidad = tk.Entry(frame, width=32, font=("Arial", 10))
        entry_cantidad.grid(row=1, column=1, pady=10)
        
        def agregar():
            producto = combo_producto.get().lower()
            
            if not producto:
                messagebox.showerror("Error", "Debe seleccionar un producto")
                return
            
            try:
                cantidad = float(entry_cantidad.get())
                
                if cantidad <= 0:
                    messagebox.showerror("Error", "La cantidad debe ser mayor a cero")
                    return
                
                wb = load_workbook(ARCHIVO_EXCEL)
                ws = wb[HOJA_INVENTARIO]
                
                producto_encontrado = False
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row, 1).value and ws.cell(row, 1).value.lower() == producto:
                        cantidad_actual = ws.cell(row, 2).value or 0
                        ws.cell(row, 2, cantidad_actual + cantidad)
                        producto_encontrado = True
                        messagebox.showinfo(
                            "Éxito",
                            f"Producto actualizado\n\n{producto.capitalize()}\nCantidad agregada: {cantidad}\nNueva cantidad: {cantidad_actual + cantidad}"
                        )
                        break
                
                if not producto_encontrado:
                    ws.append([producto, cantidad])
                    messagebox.showinfo(
                        "Éxito",
                        f"Producto agregado\n\n{producto.capitalize()}\nCantidad: {cantidad}"
                    )
                
                wb.save(ARCHIVO_EXCEL)
                ventana.destroy()
                
            except ValueError:
                messagebox.showerror("Error", "La cantidad debe ser un número válido")
            except Exception as e:
                messagebox.showerror("Error", f"Error al agregar: {str(e)}")
        
        tk.Button(
            ventana,
            text="Agregar",
            command=agregar,
            width=20,
            height=2,
            font=("Arial", 11, "bold"),
            bg="#4CAF50",
            fg="white",
            cursor="hand2"
        ).pack(pady=20)
    
    def ventana_mostrar_almacen(self):
        """Ventana para mostrar el inventario completo"""
        ventana = tk.Toplevel(self.root)
        ventana.title("Inventario Completo")
        ventana.geometry("600x400")
        
        tk.Label(
            ventana,
            text="INVENTARIO ALMACENADO",
            font=("Arial", 16, "bold")
        ).pack(pady=20)
        
        # Frame con scrollbar
        frame = tk.Frame(ventana)
        frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        scrollbar = tk.Scrollbar(frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Treeview para mostrar datos
        tree = ttk.Treeview(
            frame,
            columns=("Producto", "Cantidad"),
            show="headings",
            yscrollcommand=scrollbar.set,
            height=12
        )
        
        tree.heading("Producto", text="Producto")
        tree.heading("Cantidad", text="Cantidad")
        
        tree.column("Producto", width=300, anchor="w")
        tree.column("Cantidad", width=200, anchor="center")
        
        scrollbar.config(command=tree.yview)
        tree.pack(fill=tk.BOTH, expand=True)
        
        try:
            wb = load_workbook(ARCHIVO_EXCEL)
            ws = wb[HOJA_INVENTARIO]
            
            total_unidades = 0
            
            for row in range(2, ws.max_row + 1):
                producto = ws.cell(row, 1).value
                cantidad = ws.cell(row, 2).value or 0
                
                if producto:
                    tree.insert("", tk.END, values=(producto.capitalize(), cantidad))
                    total_unidades += cantidad
            
            tk.Label(
                ventana,
                text=f"Total de unidades: {total_unidades}",
                font=("Arial", 12, "bold")
            ).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar inventario: {str(e)}")
    
    def ventana_buscar_producto(self):
        """Ventana para buscar un producto específico"""
        ventana = tk.Toplevel(self.root)
        ventana.title("Buscar Producto")
        ventana.geometry("450x250")
        
        tk.Label(
            ventana,
            text="Buscar Producto en Almacén",
            font=("Arial", 16, "bold")
        ).pack(pady=20)
        
        tk.Label(ventana, text="Nombre del producto:", font=("Arial", 10)).pack(pady=10)
        
        entry_producto = tk.Entry(ventana, width=40, font=("Arial", 11))
        entry_producto.pack(pady=10)
        entry_producto.focus()
        
        def buscar():
            producto = entry_producto.get().strip().lower()
            
            if not producto:
                messagebox.showerror("Error", "Debe ingresar un nombre de producto")
                return
            
            try:
                wb = load_workbook(ARCHIVO_EXCEL)
                ws = wb[HOJA_INVENTARIO]
                
                for row in range(2, ws.max_row + 1):
                    prod = ws.cell(row, 1).value
                    if prod and prod.lower() == producto:
                        cantidad = ws.cell(row, 2).value or 0
                        messagebox.showinfo(
                            "Producto Encontrado",
                            f"Producto: {prod.capitalize()}\nCantidad en stock: {cantidad} unidades"
                        )
                        return
                
                messagebox.showwarning(
                    "No Encontrado",
                    f"El producto '{producto}' no se encuentra en el almacén"
                )
                
            except Exception as e:
                messagebox.showerror("Error", f"Error al buscar: {str(e)}")
        
        tk.Button(
            ventana,
            text="Buscar",
            command=buscar,
            width=20,
            height=2,
            font=("Arial", 11, "bold"),
            bg="#2196F3",
            fg="white",
            cursor="hand2"
        ).pack(pady=20)
    
    def ventana_ver_trabajos(self):
        """Ventana para ver todos los trabajos registrados"""
        ventana = tk.Toplevel(self.root)
        ventana.title("Trabajos Registrados")
        ventana.geometry("800x400")
        
        tk.Label(
            ventana,
            text="TRABAJOS REGISTRADOS",
            font=("Arial", 16, "bold")
        ).pack(pady=20)
        
        frame = tk.Frame(ventana)
        frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        scrollbar = tk.Scrollbar(frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        tree = ttk.Treeview(
            frame,
            columns=("Cliente", "Trabajo", "Fecha"),
            show="headings",
            yscrollcommand=scrollbar.set,
            height=12
        )
        
        tree.heading("Cliente", text="Cliente")
        tree.heading("Trabajo", text="Trabajo")
        tree.heading("Fecha", text="Fecha de Entrega")
        
        tree.column("Cliente", width=200, anchor="w")
        tree.column("Trabajo", width=350, anchor="w")
        tree.column("Fecha", width=150, anchor="center")
        
        scrollbar.config(command=tree.yview)
        tree.pack(fill=tk.BOTH, expand=True)
        
        try:
            wb = load_workbook(ARCHIVO_EXCEL)
            ws = wb[HOJA_TRABAJOS]
            
            total_trabajos = 0
            
            for row in range(2, ws.max_row + 1):
                cliente = ws.cell(row, 1).value
                trabajo = ws.cell(row, 2).value
                fecha = ws.cell(row, 3).value
                
                if cliente:
                    tree.insert("", tk.END, values=(cliente, trabajo.capitalize() if trabajo else "", fecha))
                    total_trabajos += 1
            
            tk.Label(
                ventana,
                text=f"Total de trabajos: {total_trabajos}",
                font=("Arial", 12, "bold")
            ).pack(pady=10)
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar trabajos: {str(e)}")
    
    def salir(self):
        """Cierra la aplicación"""
        if messagebox.askokcancel("Salir", "¿Está seguro que desea salir?"):
            self.root.quit()


def main():
    """Función principal para iniciar la interfaz gráfica"""
    root = tk.Tk()
    app = InterfazGrafica(root)
    root.mainloop()


if __name__ == "__main__":
    main()