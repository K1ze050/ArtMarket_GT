from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

# Configuración de archivos Excel
# Obtener la ruta del escritorio del usuario
ESCRITORIO = os.path.join(os.path.expanduser("~"), "Desktop")
ARCHIVO_EXCEL = os.path.join(ESCRITORIO, "base_datos.xlsx")
HOJA_INVENTARIO = "Inventario"
HOJA_TRABAJOS = "Trabajos"


def inicializar_excel():
    """Crea el archivo Excel y las hojas si no existen"""
    if not os.path.exists(ARCHIVO_EXCEL):
        wb = Workbook()
        
        # Crear hoja de Inventario
        ws_inventario = wb.active
        ws_inventario.title = HOJA_INVENTARIO
        ws_inventario.append(["Producto", "Cantidad"])
        
        # Crear hoja de Trabajos
        ws_trabajos = wb.create_sheet(HOJA_TRABAJOS)
        ws_trabajos.append(["Cliente", "Trabajo_Pendiente", "Fecha_Entrega"])
        
        wb.save(ARCHIVO_EXCEL)
        print(f"Archivo creado exitosamente en: {ARCHIVO_EXCEL}")
    else:
        print(f"Archivo ya existe en: {ARCHIVO_EXCEL}")


class Trabajo:
    # Lista de trabajos permitidos
    TRABAJOS_VALIDOS = ["corte eléctrico en vinil adhesivo", "sublimado"]
    
    def __init__(self, cliente, trabajo_pendiente, fecha_entrega):
        # Usar los setters para validación inicial
        self.cliente = cliente
        self.trabajo_pendiente = trabajo_pendiente
        self.fecha_entrega = fecha_entrega
        self.materiales_descontados = False
    
    # Property y setter para cliente (dato textual)
    @property
    def cliente(self):
        return self._cliente
    
    @cliente.setter
    def cliente(self, valor):
        try:
            if not isinstance(valor, str):
                raise ValueError("El cliente debe ser un texto")
            if len(valor.strip()) == 0:
                raise ValueError("El cliente no puede estar vacío")
            self._cliente = valor
        except Exception as e:
            print(f"Error al asignar cliente: {e}")
            self._cliente = ""
    
    # Property y setter para trabajo_pendiente (dato textual)
    @property
    def trabajo_pendiente(self):
        return self._trabajo_pendiente
    
    @trabajo_pendiente.setter
    def trabajo_pendiente(self, valor):
        try:
            if not isinstance(valor, str):
                raise ValueError("El trabajo pendiente debe ser un texto")
            
            # Convertir a minúsculas
            valor_minuscula = valor.lower().strip()
            
            if len(valor_minuscula) == 0:
                raise ValueError("El trabajo pendiente no puede estar vacío")
            
            # Validar que esté en la lista de trabajos válidos
            if valor_minuscula not in self.TRABAJOS_VALIDOS:
                raise ValueError(f"El trabajo '{valor}' no es válido. Trabajos permitidos: {', '.join(self.TRABAJOS_VALIDOS)}")
            
            self._trabajo_pendiente = valor_minuscula
        except Exception as e:
            print(f"Error al asignar trabajo pendiente: {e}")
            self._trabajo_pendiente = ""
    
    # Property y setter para fecha_entrega (formato dd-mm-yyyy)
    @property
    def fecha_entrega(self):
        return self._fecha_entrega
    
    @fecha_entrega.setter
    def fecha_entrega(self, valor):
        try:
            # Validar que sea un string
            if not isinstance(valor, str):
                raise ValueError("La fecha debe ser un texto")
            
            # Validar el formato con guiones
            if valor.count('-') != 2:
                raise ValueError("La fecha debe tener el formato dd-mm-yyyy")
            
            # Separar día, mes y año
            partes = valor.split('-')
            if len(partes[0]) != 2 or len(partes[1]) != 2 or len(partes[2]) != 4:
                raise ValueError("El formato debe ser dd-mm-yyyy (ejemplo: 25-12-2024)")
            
            # Intentar convertir el string a fecha con formato dd-mm-yyyy
            fecha_obj = datetime.strptime(valor, "%d-%m-%Y")
            self._fecha_entrega = fecha_obj
            
        except ValueError as e:
            print(f"Error al asignar fecha de entrega: {e}")
            self._fecha_entrega = None
        except Exception as e:
            print(f"Error inesperado al asignar fecha: {e}")
            self._fecha_entrega = None
    
    def descontar_materiales(self):
        """Descuenta materiales del inventario según el tipo de trabajo"""
        if not self._trabajo_pendiente:
            print("No se puede descontar materiales: trabajo no válido")
            return False
        
        if self._trabajo_pendiente == "corte eléctrico en vinil adhesivo":
            return self._descontar_vinil()
        elif self._trabajo_pendiente == "sublimado":
            return self._descontar_sublimado()
        
        return False
    
    def _descontar_vinil(self):
        """Descuenta vinil del inventario"""
        try:
            cantidad_hojas = input("\nIngrese la cantidad de hojas de vinil a usar: ")
            
            try:
                cantidad_hojas = float(cantidad_hojas)
            except ValueError:
                print("La cantidad debe ser un número válido")
                return False
            
            if cantidad_hojas < 0:
                print("No se permiten valores negativos")
                return False
            
            if cantidad_hojas == 0:
                print("La cantidad debe ser mayor a cero")
                return False
            
            # Verificar disponibilidad en Excel
            wb = load_workbook(ARCHIVO_EXCEL)
            ws = wb[HOJA_INVENTARIO]
            
            vinil_disponible = 0
            fila_vinil = None
            
            for row in range(2, ws.max_row + 1):
                producto = ws.cell(row, 1).value
                if producto and producto.lower() == "vinil":
                    vinil_disponible = ws.cell(row, 2).value or 0
                    fila_vinil = row
                    break
            
            if fila_vinil is None:
                print("El producto vinil no se encuentra en el inventario")
                return False
            
            if vinil_disponible < cantidad_hojas:
                print(f"La cantidad ingresada excede los materiales disponibles")
                print(f"Vinil disponible: {vinil_disponible}, solicitado: {cantidad_hojas}")
                return False
            
            # Descontar del inventario
            nueva_cantidad = vinil_disponible - cantidad_hojas
            ws.cell(fila_vinil, 2, nueva_cantidad)
            wb.save(ARCHIVO_EXCEL)
            
            print(f"Materiales descontados exitosamente")
            print(f"Vinil: {cantidad_hojas} hojas descontadas")
            print(f"Vinil restante: {nueva_cantidad}")
            
            self.materiales_descontados = True
            return True
            
        except Exception as e:
            print(f"Error al descontar vinil: {e}")
            return False
    
    def _descontar_sublimado(self):
        """Descuenta materiales para sublimado del inventario"""
        try:
            print("\nProductos disponibles para sublimado:")
            print("1. Playera")
            print("2. Taza")
            print("3. Vidrio")
            
            opcion = input("\nSeleccione el producto (1-3): ").strip()
            
            productos_sublimado = {
                "1": "playera",
                "2": "taza",
                "3": "vidrio"
            }
            
            if opcion not in productos_sublimado:
                print("Opción inválida")
                return False
            
            producto_elegido = productos_sublimado[opcion]
            
            cantidad = input(f"\nIngrese la cantidad de {producto_elegido}s: ")
            
            try:
                cantidad = float(cantidad)
            except ValueError:
                print("La cantidad debe ser un número válido")
                return False
            
            if cantidad < 0:
                print("No se permiten valores negativos")
                return False
            
            if cantidad == 0:
                print("La cantidad debe ser mayor a cero")
                return False
            
            # Verificar disponibilidad en Excel
            wb = load_workbook(ARCHIVO_EXCEL)
            ws = wb[HOJA_INVENTARIO]
            
            producto_disponible = 0
            papel_disponible = 0
            fila_producto = None
            fila_papel = None
            
            for row in range(2, ws.max_row + 1):
                producto = ws.cell(row, 1).value
                if producto:
                    if producto.lower() == producto_elegido:
                        producto_disponible = ws.cell(row, 2).value or 0
                        fila_producto = row
                    elif producto.lower() == "papel impresión":
                        papel_disponible = ws.cell(row, 2).value or 0
                        fila_papel = row
            
            if fila_producto is None:
                print(f"El producto {producto_elegido} no se encuentra en el inventario")
                return False
            
            if fila_papel is None:
                print("El papel impresión no se encuentra en el inventario")
                return False
            
            if producto_disponible < cantidad:
                print(f"La cantidad ingresada excede los materiales disponibles")
                print(f"{producto_elegido.capitalize()} disponible: {producto_disponible}, solicitado: {cantidad}")
                return False
            
            if papel_disponible < cantidad:
                print(f"La cantidad ingresada excede los materiales disponibles")
                print(f"Papel impresión disponible: {papel_disponible}, solicitado: {cantidad}")
                return False
            
            # Descontar del inventario
            nueva_cantidad_producto = producto_disponible - cantidad
            nueva_cantidad_papel = papel_disponible - cantidad
            
            ws.cell(fila_producto, 2, nueva_cantidad_producto)
            ws.cell(fila_papel, 2, nueva_cantidad_papel)
            wb.save(ARCHIVO_EXCEL)
            
            print(f"Materiales descontados exitosamente")
            print(f"{producto_elegido.capitalize()}: {cantidad} unidades descontadas")
            print(f"{producto_elegido.capitalize()} restante: {nueva_cantidad_producto}")
            print(f"Papel impresión: {cantidad} hojas descontadas")
            print(f"Papel impresión restante: {nueva_cantidad_papel}")
            
            self.materiales_descontados = True
            return True
            
        except Exception as e:
            print(f"Error al descontar materiales de sublimado: {e}")
            return False
    
    def guardar_en_excel(self):
        """Guarda el trabajo en Excel"""
        try:
            if self._cliente and self._trabajo_pendiente and self._fecha_entrega:
                wb = load_workbook(ARCHIVO_EXCEL)
                ws = wb[HOJA_TRABAJOS]
                
                fecha_str = self._fecha_entrega.strftime("%d-%m-%Y")
                ws.append([self._cliente, self._trabajo_pendiente, fecha_str])
                
                wb.save(ARCHIVO_EXCEL)
                print(f"Trabajo guardado en Excel exitosamente")
                return True
            else:
                print("No se puede guardar: datos inválidos")
                return False
        except Exception as e:
            print(f"Error al guardar en Excel: {e}")
            return False
    
    def imprimir_fecha(self):
        """Imprime la fecha de entrega en formato dd-mm-yyyy"""
        if self._fecha_entrega:
            print(f"Fecha de entrega: {self._fecha_entrega.strftime('%d-%m-%Y')}")
        else:
            print("Fecha de entrega: No asignada")
    
    def __str__(self):
        fecha_str = self._fecha_entrega.strftime("%d-%m-%Y") if self._fecha_entrega else "Sin fecha"
        return f"Trabajo(cliente={self.cliente}, trabajo_pendiente={self.trabajo_pendiente}, fecha_entrega={fecha_str})"


class Inventario:
    # Lista de productos permitidos
    PRODUCTOS_VALIDOS = ["playera", "taza", "papel sublimado", "vidrio", "papel impresión", "vinil"]
    
    def __init__(self, producto, cantidad_producto):
        # Inicializar bandera de producto válido
        self._producto_valido = False
        # Usar los setters para validación inicial
        self.producto = producto
        self.cantidad_producto = cantidad_producto
    
    # Property y setter para producto
    @property
    def producto(self):
        return self._producto
    
    @producto.setter
    def producto(self, valor):
        try:
            # Validar que sea un string
            if not isinstance(valor, str):
                raise ValueError("El producto debe ser un texto")
            
            # Convertir a minúsculas
            valor_minuscula = valor.lower().strip()
            
            # Validar que esté en la lista de productos válidos
            if valor_minuscula not in self.PRODUCTOS_VALIDOS:
                raise ValueError(f"El producto '{valor}' no es válido. Productos permitidos: {', '.join(self.PRODUCTOS_VALIDOS)}")
            
            # Si pasa todas las validaciones
            self._producto = valor_minuscula
            self._producto_valido = True
            
        except Exception as e:
            print(f"Error al asignar producto: {e}")
            self._producto = ""
            self._producto_valido = False
    
    # Property y setter para cantidad_producto
    @property
    def cantidad_producto(self):
        return self._cantidad_producto
    
    @cantidad_producto.setter
    def cantidad_producto(self, valor):
        try:
            # Verificar primero si el producto es válido
            if not self._producto_valido:
                raise ValueError("No se puede asignar cantidad porque el producto no es válido")
            
            # Validar que sea un número
            if not isinstance(valor, (int, float)):
                raise ValueError("La cantidad debe ser un valor numérico")
            
            # Validar que no sea negativo
            if valor < 0:
                raise ValueError("La cantidad no puede ser negativa")
            
            # Validar que no esté vacío (diferente de cero)
            if valor == 0:
                raise ValueError("La cantidad no puede ser cero")
            
            self._cantidad_producto = valor
            
            # GUARDAR EN EXCEL si todo es válido
            self._guardar_en_excel()
            
        except Exception as e:
            print(f"Error al asignar cantidad: {e}")
            self._cantidad_producto = 0
    
    def _guardar_en_excel(self):
        """Método privado para guardar el producto y cantidad en Excel"""
        try:
            if self._producto_valido and self._cantidad_producto > 0:
                wb = load_workbook(ARCHIVO_EXCEL)
                ws = wb[HOJA_INVENTARIO]
                
                # Buscar si el producto ya existe
                producto_encontrado = False
                for row in range(2, ws.max_row + 1):
                    if ws.cell(row, 1).value and ws.cell(row, 1).value.lower() == self._producto:
                        # Sumar la cantidad al producto existente
                        cantidad_actual = ws.cell(row, 2).value or 0
                        ws.cell(row, 2, cantidad_actual + self._cantidad_producto)
                        producto_encontrado = True
                        print(f"Producto {self._producto} actualizado en Excel. Nueva cantidad: {cantidad_actual + self._cantidad_producto}")
                        break
                
                # Si no existe, agregarlo
                if not producto_encontrado:
                    ws.append([self._producto, self._cantidad_producto])
                    print(f"Producto {self._producto} agregado a Excel con cantidad: {self._cantidad_producto}")
                
                wb.save(ARCHIVO_EXCEL)
        except Exception as e:
            print(f"Error al guardar en Excel: {e}")
    
    def __str__(self):
        if self._producto_valido:
            return f"Inventario(producto={self.producto}, cantidad={self.cantidad_producto})"
        else:
            return f"Inventario(producto=inválido, cantidad=no asignada)"


class Mostrar_almacen:
    """Clase para mostrar el inventario almacenado desde Excel"""
    
    def __init__(self):
        self.mostrar()
    
    def mostrar(self):
        """Muestra todo el inventario almacenado en Excel"""
        try:
            wb = load_workbook(ARCHIVO_EXCEL)
            ws = wb[HOJA_INVENTARIO]
            
            print("\nINVENTARIO ALMACENADO")
            
            # Contar productos (excluyendo encabezado)
            total_productos = ws.max_row - 1
            
            if total_productos == 0:
                print("El almacén está vacío")
            else:
                print(f"\nTotal de productos diferentes: {total_productos}")
                print("\nProductos en stock:")
                
                total_unidades = 0
                for row in range(2, ws.max_row + 1):
                    producto = ws.cell(row, 1).value
                    cantidad = ws.cell(row, 2).value or 0
                    
                    if producto:
                        print(f"{producto.capitalize()}: {cantidad} unidades")
                        total_unidades += cantidad
                
                print(f"\nTOTAL DE UNIDADES: {total_unidades}")
            
            print("")
            
        except Exception as e:
            print(f"Error al mostrar almacén: {e}")


class Buscar_en_almacen:
    """Clase para buscar productos en el almacén usando Hashing desde Excel"""
    
    def __init__(self, producto_buscar):
        self.producto_buscar = producto_buscar
        self.buscar()
    
    def buscar(self):
        """Busca un producto en el almacén de Excel usando Hashing"""
        try:
            wb = load_workbook(ARCHIVO_EXCEL)
            ws = wb[HOJA_INVENTARIO]
            
            print("\nBUSQUEDA EN ALMACEN")
            
            # Normalizar la búsqueda a minúsculas
            producto_normalizado = self.producto_buscar.lower().strip()
            
            print(f"\nBuscando: {self.producto_buscar}")
            print(f"Producto normalizado: {producto_normalizado}")
            print(f"Usando Hashing (búsqueda en Excel)...")
            
            # Crear diccionario (hash table) temporal desde Excel
            inventario_hash = {}
            for row in range(2, ws.max_row + 1):
                producto = ws.cell(row, 1).value
                cantidad = ws.cell(row, 2).value
                if producto:
                    inventario_hash[producto.lower()] = cantidad
            
            # HASHING: Acceso directo O(1)
            if producto_normalizado in inventario_hash:
                cantidad = inventario_hash[producto_normalizado]
                print(f"\nPRODUCTO ENCONTRADO")
                print(f"Producto: {producto_normalizado.capitalize()}")
                print(f"Cantidad en stock: {cantidad} unidades")
            else:
                print(f"\nPRODUCTO NO ENCONTRADO")
                print(f"El producto {self.producto_buscar} no está en el almacén")
                
                # Sugerencias de productos disponibles
                if inventario_hash:
                    print(f"\nProductos disponibles en almacén:")
                    for prod in inventario_hash.keys():
                        print(f"{prod.capitalize()}")
            
            print("")
            
        except Exception as e:
            print(f"Error al buscar en almacén: {e}")
    
    def obtener_cantidad(self):
        """Retorna la cantidad del producto buscado (si existe)"""
        try:
            wb = load_workbook(ARCHIVO_EXCEL)
            ws = wb[HOJA_INVENTARIO]
            
            producto_normalizado = self.producto_buscar.lower().strip()
            
            for row in range(2, ws.max_row + 1):
                producto = ws.cell(row, 1).value
                if producto and producto.lower() == producto_normalizado:
                    return ws.cell(row, 2).value or 0
            
            return 0
        except Exception as e:
            print(f"Error al obtener cantidad: {e}")
            return 0


# Menú principal del sistema
def menu_principal():
    """Menú interactivo para usar todas las clases del sistema"""
    
    # Inicializar archivo Excel al inicio
    inicializar_excel()
    
    while True:
        print("\nSISTEMA DE GESTION")
        print("\nSeleccione una opción:\n")
        print("1. Registrar un trabajo")
        print("2. Agregar producto al inventario")
        print("3. Mostrar almacén completo")
        print("4. Buscar producto en almacén")
        print("5. Ver trabajos registrados")
        print("6. Salir del programa")
        
        try:
            opcion = input("\nIngrese su opción (1-6): ").strip()
            
            if opcion == "1":
                registrar_trabajo()
            elif opcion == "2":
                agregar_inventario()
            elif opcion == "3":
                Mostrar_almacen()
            elif opcion == "4":
                buscar_producto()
            elif opcion == "5":
                ver_trabajos()
            elif opcion == "6":
                print("\nHasta luego\n")
                break
            else:
                print("\nOpción inválida. Por favor ingrese un número del 1 al 6.")
                input("\nPresione ENTER para continuar...")
        
        except KeyboardInterrupt:
            print("\n\nPrograma interrumpido por el usuario\n")
            break
        except Exception as e:
            print(f"\nError inesperado: {e}")
            input("\nPresione ENTER para continuar...")


def registrar_trabajo():
    """Función para registrar un nuevo trabajo"""
    print("\nREGISTRAR TRABAJO")
    
    print("\nTrabajos válidos:")
    for i, trab in enumerate(Trabajo.TRABAJOS_VALIDOS, 1):
        print(f"{i}. {trab.capitalize()}")
    
    try:
        cliente = input("\nIngrese el nombre del cliente: ")
        trabajo_pendiente = input("Ingrese el tipo de trabajo: ")
        fecha_entrega = input("Ingrese la fecha de entrega (dd-mm-yyyy): ")
        
        trabajo = Trabajo(cliente=cliente, trabajo_pendiente=trabajo_pendiente, fecha_entrega=fecha_entrega)
        
        if not trabajo._trabajo_pendiente:
            print("\nNo se puede continuar sin un trabajo válido")
            input("\nPresione ENTER para volver al menú...")
            return
        
        print("\nTrabajo registrado:")
        print(f"{trabajo}")
        trabajo.imprimir_fecha()
        
        # Descontar materiales
        print("\nProcediendo a descontar materiales del inventario...")
        if trabajo.descontar_materiales():
            # Solo guardar en Excel si se descontaron los materiales exitosamente
            trabajo.guardar_en_excel()
        else:
            print("\nEl trabajo no se guardará porque no se pudieron descontar los materiales")
        
    except Exception as e:
        print(f"\nError al registrar trabajo: {e}")
    
    input("\nPresione ENTER para volver al menú...")


def agregar_inventario():
    """Función para agregar productos al inventario"""
    print("\nAGREGAR AL INVENTARIO")
    
    print("\nProductos válidos:")
    for i, prod in enumerate(Inventario.PRODUCTOS_VALIDOS, 1):
        print(f"{i}. {prod.capitalize()}")
    
    try:
        producto = input("\nIngrese el nombre del producto: ")
        cantidad = input("Ingrese la cantidad: ")
        
        # Intentar convertir la cantidad a número
        try:
            cantidad = float(cantidad)
        except ValueError:
            print("\nLa cantidad debe ser un número válido")
            input("\nPresione ENTER para volver al menú...")
            return
        
        inventario = Inventario(producto=producto, cantidad_producto=cantidad)
        
    except Exception as e:
        print(f"\nError al agregar inventario: {e}")
    
    input("\nPresione ENTER para volver al menú...")


def buscar_producto():
    """Función para buscar un producto en el almacén"""
    print("\nBUSCAR PRODUCTO")
    
    try:
        producto = input("\nIngrese el nombre del producto a buscar: ")
        
        if producto.strip() == "":
            print("\nDebe ingresar un nombre de producto")
        else:
            Buscar_en_almacen(producto)
        
    except Exception as e:
        print(f"\nError al buscar producto: {e}")
    
    input("\nPresione ENTER para volver al menú...")


def ver_trabajos():
    """Función para ver todos los trabajos registrados"""
    print("\nTRABAJOS REGISTRADOS")
    
    try:
        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb[HOJA_TRABAJOS]
        
        total_trabajos = ws.max_row - 1
        
        if total_trabajos == 0:
            print("\nNo hay trabajos registrados")
        else:
            print(f"\nTotal de trabajos: {total_trabajos}\n")
            
            for row in range(2, ws.max_row + 1):
                cliente = ws.cell(row, 1).value
                trabajo = ws.cell(row, 2).value
                fecha = ws.cell(row, 3).value
                
                print(f"{row-1}. Cliente: {cliente}")
                print(f"   Trabajo: {trabajo}")
                print(f"   Fecha: {fecha}\n")
        
    except Exception as e:
        print(f"\nError al mostrar trabajos: {e}")
    
    input("\nPresione ENTER para volver al menú...")

menu_principal()